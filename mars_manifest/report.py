"""Report rendering: markdown and xlsx (HANDOFF.md §5, M5).

Strictly a rendering layer — every number comes from an engine result.
The xlsx budget workbook mirrors ../Mars-Precursor-Engineering-Budget.xlsx
(Assumptions / Line Items / Power & Storage / Summary) with added
Per-Ship and Launch & Cost sheets plus charts.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .budgets import BudgetResult
from .campaign import CampaignResult
from .catalog import Catalog
from .models import Assumptions, Mission
from .packing import PackingResult

_HDR = Font(bold=True)


# ---------------------------------------------------------------------------
# markdown
# ---------------------------------------------------------------------------

def _md_table(headers: list[str], rows: list[list]) -> str:
    def fmt(v):
        if isinstance(v, float):
            return f"{v:,.2f}" if abs(v) < 100 else f"{v:,.1f}"
        return str(v)
    out = ["| " + " | ".join(headers) + " |", "|" + "|".join("---" for _ in headers) + "|"]
    out += ["| " + " | ".join(fmt(v) for v in row) + " |" for row in rows]
    return "\n".join(out)


def budget_markdown(budget: BudgetResult, mission: Mission,
                    packing: Optional[PackingResult] = None) -> str:
    b = budget
    parts = [f"# Budget — {mission.id}", ""]
    if mission.objective:
        parts += [f"*{mission.objective}*", ""]
    parts += [f"Power path: **{b.power_path}**", ""]

    parts += ["## Surface load", _md_table(
        ["Metric", "Value"],
        [["Average load (kW)", b.loads.avg_kw],
         ["Connected peak (kW)", b.loads.connected_peak_kw],
         ["Simultaneous peak (kW)", b.loads.simultaneous_peak_kw],
         ["Daily energy (kWh)", b.loads.daily_energy_kwh]]), ""]

    parts += ["## Power options", _md_table(
        ["Option", "Generation", "Mass (t)", "Storage (t)", "Note"],
        [["Solar + battery", f"{b.solar.array_m2:,.0f} m2", b.solar.array_mass_t,
          b.solar.night_battery_t,
          f"storm battery {b.solar.dust_storm_battery_critical_t:,.0f} t survival-only "
          f"({b.solar.dust_storm_battery_t:,.0f} t if production rides through)"
          + (" -- INFEASIBLE" if b.solar.dust_storm_infeasible else "")],
         ["Fission + buffer", f"{b.fission.units} units", b.fission.mass_t,
          b.fission.buffer_battery_t, "weather-independent"]]), ""]

    parts += ["## Mass budget", _md_table(
        ["Line", "Mass (t)"],
        [["Fixed hardware", b.mass.fixed_hardware_t],
         ["Generation", b.mass.generation_t],
         ["Storage", b.mass.storage_t],
         ["Spares / redundancy", b.mass.spares_t],
         ["Contingency / harness", b.mass.contingency_t],
         ["(memo) consumables in fixed", b.mass.consumables_t],
         ["**GRAND TOTAL**", b.mass.grand_total_t]]), ""]

    parts += ["## Volume budget", _md_table(
        ["Line", "Volume (m3)"],
        [["Fixed hardware", b.volume.fixed_m3],
         ["Generation", b.volume.generation_m3],
         ["Storage", b.volume.storage_m3],
         ["Spares (est)", b.volume.spares_m3],
         ["Raw stowed", b.volume.raw_m3],
         ["Effective (with packing)", b.volume.effective_m3]]), ""]

    parts += ["## Cargo hardware cost", _md_table(
        ["Line", "Low ($M)", "High ($M)"],
        [["Fixed hardware", b.cost.fixed_hardware_low_musd, b.cost.fixed_hardware_high_musd],
         ["Generation", b.cost.generation_low_musd, b.cost.generation_high_musd],
         ["Storage", b.cost.storage_low_musd, b.cost.storage_high_musd],
         ["**Total cargo**", b.cost.cargo_low_musd, b.cost.cargo_high_musd]]), ""]

    if b.capacity:
        c = b.capacity
        parts += ["## Capacity check", _md_table(
            ["Metric", "Value"],
            [["Ships in batch", c.ships],
             ["Mass capacity (t)", c.mass_capacity_t],
             ["Mass utilisation", f"{c.mass_utilisation:.1%}"],
             ["Mass margin (t)", c.mass_margin_t],
             ["Volume capacity (m3)", c.volume_capacity_m3],
             ["Volume utilisation", f"{c.volume_utilisation:.1%}"]]), ""]

    if packing:
        parts += [packing_markdown(packing), ""]

    if b.warnings:
        parts += ["## Warnings"] + [f"- {w}" for w in b.warnings] + [""]
    return "\n".join(parts)


def packing_markdown(packing: PackingResult) -> str:
    rows = [[s.index, s.mass_t, s.volume_m3, f"{s.mass_utilisation:.0%}",
             f"{s.volume_utilisation:.0%}", s.binding_constraint,
             ", ".join(f"{cid} x{qty:g}" for cid, qty in s.items)]
            for s in packing.ships]
    lm = packing.launch
    parts = [f"## Ship packing ({packing.policy})", _md_table(
        ["Ship", "Mass (t)", "Vol (m3)", "Mass util", "Vol util", "Binding", "Items"], rows), "",
        "## Launch math", _md_table(
            ["Metric", "Value"],
            [["Cargo ship launches", lm.cargo_ship_launches],
             ["Tankers per ship", lm.tankers_per_ship],
             ["Tanker launches", lm.tanker_launches],
             ["**Total launches**", lm.total_launches]]
            + ([["Cargo-ship rate (expended)", f"${lm.cargo_ship_rate_musd:g}M/launch"],
                ["Tanker rate (reused)", f"${lm.tanker_rate_musd:g}M/launch"]]
               if lm.split_rates else
               [[f"Rate @ {lm.launch_cost_tier}", f"${lm.per_launch_cost_musd:g}M/launch"]])
            + [["**Launch campaign cost**", f"${lm.launch_cost_musd:,.0f}M"]])]
    if packing.single_points:
        parts += ["", "## Single points of failure (one ship carries the whole capability)"]
        parts += [f"- {cid} — sole carrier of gate `{gate}`" for cid, gate in packing.single_points]
    return "\n".join(parts)


def chain_design_markdown(d) -> str:
    parts = ["# Rate-matched chain design",
             f"Target: {d.target_rate_kg_hr:,.1f} kg/hr nameplate "
             f"(~{d.target_tonnes_per_synod:,.0f} t per synod at plant availability)", ""]
    rows = [[s.key, s.component_id, f"{s.unit_rate_kg_hr:,.1f}", s.units_required,
             f"{s.utilization:.0%}", s.mass_t, s.avg_kw] for s in d.steps]
    parts += [_md_table(["Step", "Component", "Unit rate (kg/hr)", "Units", "Utilization",
                         "Mass (t)", "Avg kW"], rows), ""]
    parts += [_md_table(["Rollup", "Value"], [
        ["Chain hardware", f"{d.chain_mass_t:,.1f} t / {d.chain_avg_kw:,.0f} kW avg"],
        ["Fission required", f"{d.fission_units} units / {d.fission_mass_t:,.0f} t"],
        ["Buffer battery", f"{d.buffer_battery_t:,.1f} t"],
        ["**Total plant**", f"{d.total_mass_t:,.1f} t"],
        ["Hardware cost", f"${d.cost_low_musd:,.0f}M-${d.cost_high_musd:,.0f}M"]]), ""]
    if d.notes:
        parts += ["## Granularity slack"] + [f"- {n}" for n in d.notes] + [""]
    return "\n".join(parts)


def requirements_markdown(matrix, requirements, campaign_id: str) -> str:
    parts = [f"# Requirements buy-off matrix — {campaign_id}", "",
             "Status vs plan: CLOSED (on plan) / EARLY / LATE / OPEN; "
             "recurring requirements: PASS / FAIL across all windows.", ""]

    counts = " · ".join(f"{k} {v}" for k, v in sorted(matrix.counts.items()))
    parts += [f"**{counts}**", ""]

    parts += ["## Buy-off by flight window"]
    for wid, ids in matrix.by_window.items():
        parts.append(f"- **{wid}**: {', '.join(ids) if ids else '—'}")
    parts.append("")

    by_id = {r.id: r for r in requirements}
    rows = []
    for v in matrix.verdicts:
        r = v.requirement
        indent = "&nbsp;&nbsp;" * r.level
        rows.append([
            f"{indent}{r.id}", r.statement, r.method,
            r.planned_window or ("every window" if r.recurring else "rollup"),
            v.closed_window or ("; ".join(v.failed_windows) if v.failed_windows else "—"),
            v.status,
        ])
    parts += ["## Full matrix", _md_table(
        ["Req", "Statement", "Verify", "Planned", "Actual", "Status"], rows), ""]

    if matrix.open_ids:
        parts += ["## Open items"] + [f"- {rid}: {by_id[rid].statement}" for rid in matrix.open_ids] + [""]
    return "\n".join(parts)


def edl_markdown(result) -> str:
    parts = ["# EDL risk — landing probability & demonstrated reliability", "",
             "Per-ship success improves as EDL matures; reliability is the "
             "rule-of-three floor (0 failures in N landings → failure rate "
             "≤ ~3/N at 95% confidence). Assumes independent ship losses.", ""]
    rows = [[w.window_id, w.ships, f"{w.edl_success_prob:.0%}",
             f"{w.expected_ships_landed:,.1f}",
             f"{w.demonstrated_reliability:.1%}"] for w in result.windows]
    parts += [_md_table(
        ["Window", "Ships", "Per-ship P(land)", "Expected landed",
         "Demonstrated reliability"], rows), ""]
    # first window with population that also clears a useful reliability bar
    crewed = [w for w in result.windows if any(m.crewed and not m.blocked for m in w.missions)]
    if crewed:
        c = crewed[0]
        parts += [f"First crew ({c.window_id}) commits at "
                  f"{c.demonstrated_reliability:.0%} demonstrated EDL reliability "
                  f"after the prior windows' landings.", ""]
    return "\n".join(parts)


def lifecycle_markdown(rep) -> str:
    parts = ["# Lifecycle review — risk buy-down & idle hardware", ""]
    rows = [[p.window_id, ", ".join(p.retired) or "-", f"{p.weight_retired:.0f}",
             f"{p.cumulative_fraction:.0%}"] for p in rep.risk_curve]
    parts += ["## Risk buy-down curve", _md_table(
        ["Window", "Gates retired", "Weight", "Cumulative risk retired"], rows), ""]
    if rep.idle_items:
        rows = [[i.component_id, i.delivered_window, f"{i.qty:g}", i.mass_t,
                 i.idle_synods, i.tonne_years,
                 f"{i.demo_exempt_qty:g}" if i.demo_exempt_qty else "",
                 "SHELF LIFE" if i.shelf_life_flag else ""] for i in rep.idle_items]
        parts += ["## Crew-era hardware timing", _md_table(
            ["Component", "Delivered", "Qty", "Mass (t)", "Idle synods",
             "Idle t-years", "Demo exempt", "Flag"], rows), "",
            f"Total idle: **{rep.idle_tonne_years:,.0f} tonne-years** beyond the "
            f"one-window verification margin.", ""]
    if rep.findings:
        parts += ["## Findings"] + [f"- {f}" for f in rep.findings] + [""]
    return "\n".join(parts)


def loss_markdown(lt) -> str:
    if lt.tolerant:
        return ("## Loss tolerance\n"
                "TOLERANT — no single-ship EDL loss costs any capability.")
    parts = ["## Loss tolerance", "NOT TOLERANT — single-ship losses that kill capabilities:"]
    for ship, lost in lt.vulnerable_ships:
        parts.append(f"- Ship {ship}: loses {', '.join(lost)}")
    return "\n".join(parts)


def isru_markdown(r) -> str:
    title = "# ISRU propellant chain"
    if getattr(r, "mode", "sabatier") == "oxygen_only":
        title += " — OXYGEN-ONLY FALLBACK (CH4 from Earth)"
    elif getattr(r, "mode", "sabatier") == "h2_import":
        title += " — H2-IMPORT MODE (LH2 from Earth, no water mining)"
    parts = [title, ""]
    rows = [[("-> " if s.is_bottleneck else "") + s.key, s.component_id, f"{s.units:g}",
             s.avg_kw, f"{s.throughput_kg_hr:,.1f} kg/hr {s.commodity}",
             s.propellant_rate_kg_hr] for s in r.steps]
    parts += [_md_table(["Step", "Component", "Units", "Avg kW", "Throughput",
                         "Prop-equiv kg/hr"], rows), ""]
    parts += ["## Production & energy", _md_table(
        ["Metric", "Value"],
        [["Chain rate (bottleneck: " + r.bottleneck + ")", f"{r.propellant_rate_kg_hr:,.1f} kg/hr"],
         ["Propellant per sol", f"{r.propellant_kg_per_sol:,.0f} kg"],
         ["Per production window", f"{r.tonnes_per_window:,.1f} t"],
         ["One return load", f"{r.return_load_t:,.0f} t -> {r.fraction_of_return_load:.1%} per window"],
         ["Time to one return load", f"{r.sols_to_return_load:,.0f} sols (~{r.years_to_return_load:,.1f} yr)"
          if r.sols_to_return_load else "n/a"],
         ["Net water demand", f"{r.net_water_kg_per_sol:,.0f} kg/sol; {r.water_for_return_load_t:,.0f} t per load"],
         ["Specific energy", f"{r.spec_energy_kwh_per_kg:,.2f} kWh/kg propellant"],
         ["Energy per return load", f"{r.energy_per_return_load_gwh:,.2f} GWh"],
         ["Full-scale power (one load per window)", f"{r.full_scale_kw_required:,.0f} kW continuous"],
         ["O2 balance", f"{r.o2_surplus_per_kg_ch4:+.2f} kg O2 per kg CH4 vs engine demand"]]
        + ([["CH4 imported from Earth",
             f"{r.ch4_import_t_per_load:,.0f} t per load (~{r.ch4_import_ships_per_load} extra ships)"]]
           if getattr(r, "mode", "sabatier") == "oxygen_only" else [])
        + ([["LH2 imported from Earth",
             f"{r.h2_import_t_per_load:,.0f} t per load (~{r.h2_import_ships_per_load} ships, volume-bound)"],
            ["Water mined", "none (recycled from Sabatier product)"]]
           if getattr(r, "mode", "sabatier") == "h2_import" else [])), ""]
    if r.warnings:
        parts += ["## Findings"] + [f"- {w}" for w in r.warnings] + [""]
    return "\n".join(parts)


# Layman-conversion anchors (rendering only — never used in engine math).
# Sources: US EIA average household ~1.2 kW continuous; NASA BVAD minimum
# acceptable net habitable volume ~25 m3/person for long-duration; open-loop
# consumables ~5 kg/person-day (food+water+O2 with partial recycling);
# ISS assembled mass ~420 t.
_HOME_KW = 1.2
_HABITABLE_M3_PER_PERSON = 25.0
_CONSUMABLES_KG_PER_PERSON_DAY = 5.0
_ISS_MASS_T = 420.0


def window_enablements(w, catalog: Catalog, a: Assumptions,
                       next_label: Optional[str] = None) -> list[str]:
    """Plain-English 'what this window buys us', derived from engine outputs.

    Propellant figures are cumulative production by the time the NEXT window
    arrives (production runs during the synod after this window's landings);
    pass next_label so the text can say so explicitly.
    """
    out: list[str] = []
    inv = {cid: qty for cid, qty, _ in w.surface_inventory}
    load_t = a.get("isru.return_propellant_t")
    avail = a.get("isru.plant_availability", 1.0)
    sol_h = a.get("power.sol_hours")
    crewed = any(m.crewed and not m.blocked for m in w.missions)
    by_next = f"by the {next_label} window" if next_label else "by the next window's arrival"

    if crewed:
        out.append("First crew lands at a base that is already powered, provisioned, "
                   "and holds their return propellant.")
    if w.population > 0:
        out.append(f"Resident population: {w.population:,} — every per-capita clock "
                   f"(power, food area, habitat volume, imports) now has a denominator.")
    agri = inv.get("agri_module", 0)
    if agri and w.population > 0:
        fed = agri * 4  # 200 m2/module at the 50 m2/person planning value
        out.append(f"Agriculture: {agri:,.0f} modules can feed ~{fed:,.0f} people "
                   f"({'surplus' if fed >= w.population else 'DEFICIT'} vs "
                   f"{w.population:,} residents).")
    if w.import_required_t > 0:
        out.append(f"Recurring imports: residents need {w.import_required_t:,.0f} t this "
                   f"synod at {w.import_rate_t_py:g} t/person/yr (closure stage: "
                   f"{w.closure_stage.replace('_', ' ')}); manifests deliver "
                   f"{w.import_delivered_t:,.0f} t.")

    kwe = w.installed_generation_kwe
    if kwe > 0:
        out.append(f"{kwe:,.0f} kWe of storm-proof power — roughly "
                   f"{kwe / _HOME_KW:,.0f} homes' worth, with "
                   f"{kwe - w.surface_avg_load_kw:,.0f} kW of headroom over the installed load.")
    if w.installed_storage_kwh > 0 and w.surface_avg_load_kw > 0:
        hours = w.installed_storage_kwh / w.surface_avg_load_kw
        out.append(f"Batteries hold {w.installed_storage_kwh / 1000:,.1f} MWh — "
                   f"{hours:,.0f} hours of full-base power with zero generation "
                   f"(dust storms are a non-event on fission).")

    kg_day = w.isru_rate_kg_hr * sol_h * avail
    if kg_day > 0:
        days_per_load = load_t * 1000.0 / kg_day
        if days_per_load > 1200:
            pace = f"a full Starship return load every {days_per_load / 365.25:,.1f} years (pilot proof, not production)"
        else:
            pace = f"a full Starship return load every {days_per_load / 30.4:,.0f} months"
        out.append(f"Fuel plant produces ~{kg_day / 1000:,.1f} t of methalox per day — {pace}.")
    if w.propellant_cumulative_t > 0:
        loads = w.propellant_cumulative_t / load_t
        line = (f"{w.propellant_cumulative_t:,.0f} t of propellant banked {by_next} — "
                f"{loads:,.1f} Starship return flights' worth.")
        if crewed:
            line += f" The crew's ride home could be filled {loads:,.1f} times over."
        out.append(line)

    # habitat_module carries deployed volume in the catalog; the inflatable
    # deploys ~300 m3 from its 75 m3 stowed figure
    hab_qty = inv.get("habitat_module", 0) + inv.get("habitat_inflatable", 0)
    if hab_qty:
        m3 = hab_qty * 300.0
        out.append(f"{m3:,.0f} m³ of pressurized volume beyond the ships — long-duration "
                   f"habitable space for ~{m3 / _HABITABLE_M3_PER_PERSON:,.0f} people.")
    cache_qty = inv.get("consumables_cache", 0)
    if cache_qty:
        cache_t = cache_qty * catalog.get("consumables_cache").unit_mass_t
        person_days = cache_t * 1000.0 / _CONSUMABLES_KG_PER_PERSON_DAY
        out.append(f"{cache_t:,.0f} t of cached food/water/O2 — a crew of 12 could live on "
                   f"it for {person_days / 12 / 30.4:,.0f} months with zero recycling.")

    robots = inv.get("optimus_robot", 0)
    rovers = inv.get("rover_unpressurized", 0)
    if robots or rovers:
        out.append(f"Robot workforce: {robots:,.0f} humanoids and {rovers:,.0f} rovers"
                   + ("." if crewed else " — the only crew on the planet."))

    out.append(f"{w.surface_hardware_t:,.0f} t of hardware on the surface — "
               f"{w.surface_hardware_t / _ISS_MASS_T:,.1f}× the mass of the ISS, and "
               f"{w.surface_hardware_t / 1_000_000 * 100:.2f}% of the ~1,000,000 t "
               f"Musk estimates for a self-sufficient city.")
    return out


def campaign_markdown(result: CampaignResult, catalog: Optional[Catalog] = None,
                      assumptions: Optional[Assumptions] = None) -> str:
    parts = [f"# Campaign — {result.campaign_id}", ""]
    rows = []
    for w in result.windows:
        crew = any(m.crewed and not m.blocked for m in w.missions)
        blocked = any(m.blocked for m in w.missions)
        status = "CREWED" if crew else ("BLOCKED" if blocked else "robotic")
        rows.append([w.window_id, w.objective, w.ships, w.total_launches,
                     w.mass_delivered_t, f"{w.installed_generation_kwe:,.0f}",
                     f"{w.propellant_cumulative_t:,.0f}",
                     f"${w.launch_cost_musd:,.0f}M", status,
                     ", ".join(w.new_capabilities) or "-"])
    parts += [_md_table(
        ["Window", "Objective", "Ships", "Launches", "Mass (t)", "Power (kWe)",
         "Propellant banked by next window (t)", "Launch cost", "Status",
         "New capabilities"], rows), ""]

    c = result.cumulative
    parts += ["## Cumulative", _md_table(
        ["Metric", "Value"],
        [["Ships landed", c["ships"]],
         ["Total launches", c["total_launches"]],
         ["Mass delivered (t)", c["mass_delivered_t"]],
         ["Launch cost ($M)", c["launch_cost_musd"]],
         ["Cargo cost ($M)", f"{c['cargo_cost_low_musd']:,.0f}-{c['cargo_cost_high_musd']:,.0f}"],
         ["First crew window", result.first_crew_window or "not achieved"]]), ""]

    if catalog is not None and assumptions is not None:
        parts += ["## What each window enables"]
        for i, w in enumerate(result.windows):
            nxt = result.windows[i + 1].window_id if i + 1 < len(result.windows) else None
            parts.append(f"### {w.window_id}")
            parts += [f"- {line}" for line in window_enablements(w, catalog, assumptions, nxt)]
        parts.append("")

    if result.violations:
        parts += ["## Gating violations"] + [f"- {v}" for v in result.violations] + [""]
    all_warnings = [w for wr in result.windows for w in wr.warnings]
    if all_warnings:
        parts += ["## Advisories"] + [f"- {w}" for w in all_warnings] + [""]
    return "\n".join(parts)


def comparison_markdown(comp) -> str:
    parts = [f"# Scenario comparison — {comp.scenario_a} vs {comp.scenario_b}", ""]
    rows = [[r.metric, r.a, r.b, "*" if r.changed else ""] for r in comp.rows]
    parts += [_md_table(["Metric", comp.scenario_a, comp.scenario_b, "diff"], rows), ""]
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# xlsx
# ---------------------------------------------------------------------------

def _sheet_rows(ws, rows: list[list], header_rows: int = 1) -> None:
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r <= header_rows:
                cell.font = _HDR
    for c in range(1, (max(len(r) for r in rows) if rows else 1) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24


def budget_xlsx(budget: BudgetResult, mission: Mission, catalog: Catalog,
                assumptions: Assumptions, path: str | Path,
                packing: Optional[PackingResult] = None) -> Path:
    b = budget
    wb = Workbook()

    ws = wb.active
    ws.title = "Assumptions"
    a = assumptions
    _sheet_rows(ws, [
        ["Assumption", "Value", "Unit"],
        ["Scenario", a.name, ""],
        ["Payload mass per ship", a.get("fleet.payload_mass_per_ship_t"), "t"],
        ["Payload volume per ship", a.get("fleet.payload_volume_per_ship_m3"), "m3"],
        ["Packing / stowage efficiency", a.get("fleet.packing_efficiency"), "frac"],
        ["Tankers per ship", a.get("fleet.tankers_per_ship"), "launches"],
        ["Martian sol length", a.get("power.sol_hours"), "hr"],
        ["Night duration", a.get("power.night_hours"), "hr"],
        ["Dust-storm autonomy target", a.get("power.dust_storm_autonomy_days"), "days"],
        ["Peak diversity factor", a.get("power.diversity_factor"), "frac"],
        ["Solar 24h-average yield", a.get("power.solar_yield_w_per_m2"), "W/m2"],
        ["Solar specific mass", a.get("power.solar_kg_per_m2"), "kg/m2"],
        ["Battery specific energy", a.get("power.battery_wh_per_kg"), "Wh/kg"],
        ["Battery energy density", a.get("power.battery_kwh_per_m3"), "kWh/m3"],
        ["Fission unit output", a.get("power.fission_unit_kwe"), "kWe"],
        ["Fission unit mass", a.get("power.fission_unit_mass_t"), "t"],
        ["Fission unit stowed volume", a.get("power.fission_unit_volume_m3"), "m3"],
        ["Fission buffer storage", a.get("power.fission_buffer_hours"), "hr"],
        ["Spares / redundancy", a.get("overheads.spares_fraction_of_dry"), "frac of dry"],
        ["Contingency / harness", a.get("overheads.contingency_fraction"), "frac"],
        ["Power path", b.power_path, ""],
    ])

    ws = wb.create_sheet("Line Items")
    rows = [["Group", "Item", "Ship", "Unit mass (t)", "Qty", "Total mass (t)",
             "Unit vol (m3)", "Total vol (m3)", "Peak/unit (kW)", "Duty",
             "Total peak (kW)", "Avg power (kW)"]]
    for item in mission.manifest:
        comp = catalog.get(item.component_id)
        rows.append([comp.group, comp.name, item.ship or "", comp.unit_mass_t, item.qty,
                     comp.unit_mass_t * item.qty, comp.unit_volume_m3,
                     comp.unit_volume_m3 * item.qty, comp.peak_power_kw, comp.duty_cycle,
                     comp.peak_power_kw * item.qty,
                     comp.peak_power_kw * item.qty * comp.duty_cycle])
    rows.append(["", "TOTAL (fixed hardware)", "", "", "", b.mass.fixed_hardware_t, "",
                 b.volume.fixed_m3, "", "", b.loads.connected_peak_kw, b.loads.avg_kw])
    _sheet_rows(ws, rows)

    ws = wb.create_sheet("Power & Storage")
    _sheet_rows(ws, [
        ["Metric", "Value", "Unit"],
        ["Total average load", b.loads.avg_kw, "kW"],
        ["Total connected peak", b.loads.connected_peak_kw, "kW"],
        ["Simultaneous peak", b.loads.simultaneous_peak_kw, "kW"],
        ["Daily energy demand", b.loads.daily_energy_kwh, "kWh"],
        ["", "", ""],
        ["OPTION A — SOLAR + BATTERY", "", ""],
        ["Array area required", b.solar.array_m2, "m2"],
        ["Solar array mass", b.solar.array_mass_t, "t"],
        ["Night storage energy", b.solar.night_kwh, "kWh"],
        ["Night battery mass", b.solar.night_battery_t, "t"],
        ["Dust-storm storage energy", b.solar.dust_storm_kwh, "kWh"],
        ["Dust-storm battery mass", b.solar.dust_storm_battery_t, "t"],
        ["Solar-only feasible?", "NO" if b.solar.dust_storm_infeasible else "yes", ""],
        ["", "", ""],
        ["OPTION B — FISSION + BUFFER", "", ""],
        ["Fission units required", b.fission.units, "units"],
        ["Fission mass", b.fission.mass_t, "t"],
        ["Fission stowed volume", b.fission.volume_m3, "m3"],
        ["Buffer storage energy", b.fission.buffer_kwh, "kWh"],
        ["Buffer battery mass", b.fission.buffer_battery_t, "t"],
        ["Buffer battery volume", b.fission.buffer_battery_m3, "m3"],
    ])
    chart = BarChart()
    chart.title = "Storage mass: solar vs fission (t)"
    chart.type = "col"
    ws2 = wb.create_sheet("_chartdata")
    _sheet_rows(ws2, [["Option", "Storage mass (t)"],
                      ["Solar night", b.solar.night_battery_t],
                      ["Solar dust-storm", b.solar.dust_storm_battery_t],
                      ["Fission buffer", b.fission.buffer_battery_t]])
    data = Reference(ws2, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")
    ws2.sheet_state = "hidden"

    ws = wb.create_sheet("Summary")
    rows = [["MASS BUDGET", "Mass (t)"]]
    rows += [[grp, val] for grp, val in b.mass.by_group.items()]
    rows += [["Subtotal — fixed hardware", b.mass.fixed_hardware_t],
             [f"Generation ({b.power_path})", b.mass.generation_t],
             ["Storage", b.mass.storage_t],
             ["(memo) consumables in subtotal", b.mass.consumables_t],
             ["Spares / redundancy", b.mass.spares_t],
             ["Contingency / harness", b.mass.contingency_t],
             ["GRAND TOTAL MASS", b.mass.grand_total_t],
             ["", ""],
             ["VOLUME BUDGET", "Vol (m3)"],
             ["Fixed hardware volume", b.volume.fixed_m3],
             ["Generation volume", b.volume.generation_m3],
             ["Storage volume", b.volume.storage_m3],
             ["Spares volume (est)", b.volume.spares_m3],
             ["Raw stowed volume", b.volume.raw_m3],
             ["Effective (with packing)", b.volume.effective_m3],
             ["", ""],
             ["CARGO COST", "Low ($M)"],
             ["Fixed hardware", b.cost.fixed_hardware_low_musd],
             ["Generation", b.cost.generation_low_musd],
             ["Storage", b.cost.storage_low_musd],
             ["Total cargo (low)", b.cost.cargo_low_musd],
             ["Total cargo (high)", b.cost.cargo_high_musd]]
    if b.capacity:
        rows += [["", ""],
                 ["Batch mass capacity (t)", b.capacity.mass_capacity_t],
                 ["Mass utilisation", b.capacity.mass_utilisation],
                 ["Mass margin (t)", b.capacity.mass_margin_t],
                 ["Batch volume capacity (m3)", b.capacity.volume_capacity_m3],
                 ["Volume utilisation", b.capacity.volume_utilisation]]
    _sheet_rows(ws, rows)

    if packing:
        ws = wb.create_sheet("Per-Ship")
        rows = [["Ship", "Mass (t)", "Vol (m3)", "Mass util", "Vol util", "Binding", "Items"]]
        for s in packing.ships:
            rows.append([s.index, s.mass_t, s.volume_m3, s.mass_utilisation,
                         s.volume_utilisation, s.binding_constraint,
                         ", ".join(f"{cid} x{qty:g}" for cid, qty in s.items)])
        _sheet_rows(ws, rows)
        chart = BarChart()
        chart.title = "Per-ship mass (t)"
        chart.type = "col"
        data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(packing.ships))
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(packing.ships))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "I2")

        ws = wb.create_sheet("Launch & Cost")
        lm = packing.launch
        _sheet_rows(ws, [
            ["Metric", "Value"],
            ["Cargo ship launches", lm.cargo_ship_launches],
            ["Tankers per ship", lm.tankers_per_ship],
            ["Tanker launches", lm.tanker_launches],
            ["Total launches", lm.total_launches],
            ["Launch cost tier", lm.launch_cost_tier],
            ["Per-launch cost ($M)", lm.per_launch_cost_musd],
            ["Launch campaign cost ($M)", lm.launch_cost_musd],
            ["Cargo hardware low ($M)", b.cost.cargo_low_musd],
            ["Cargo hardware high ($M)", b.cost.cargo_high_musd],
        ])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def campaign_xlsx(result: CampaignResult, path: str | Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Windows"
    rows = [["Window", "Objective", "Ships", "Launches", "Mass delivered (t)",
             "Installed power (kWe)", "Launch cost ($M)", "Cargo low ($M)",
             "Cargo high ($M)", "New capabilities"]]
    for w in result.windows:
        rows.append([w.window_id, w.objective, w.ships, w.total_launches,
                     w.mass_delivered_t, w.installed_generation_kwe,
                     w.launch_cost_musd, w.cargo_cost_low_musd, w.cargo_cost_high_musd,
                     ", ".join(w.new_capabilities)])
    _sheet_rows(ws, rows)
    chart = BarChart()
    chart.title = "Mass delivered per window (t)"
    chart.type = "col"
    data = Reference(ws, min_col=5, min_row=1, max_row=1 + len(result.windows))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(result.windows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "L2")

    ws = wb.create_sheet("Cumulative")
    c = result.cumulative
    _sheet_rows(ws, [
        ["Metric", "Value"],
        ["Ships landed", c["ships"]],
        ["Total launches", c["total_launches"]],
        ["Mass delivered (t)", c["mass_delivered_t"]],
        ["Launch cost ($M)", c["launch_cost_musd"]],
        ["Cargo cost low ($M)", c["cargo_cost_low_musd"]],
        ["Cargo cost high ($M)", c["cargo_cost_high_musd"]],
        ["First crew window", result.first_crew_window or "not achieved"],
    ])

    ws = wb.create_sheet("Capabilities")
    rows = [["Window", "Capabilities unlocked (cumulative)"]]
    for w in result.windows:
        rows.append([w.window_id, ", ".join(w.capabilities_after)])
    _sheet_rows(ws, rows)

    ws = wb.create_sheet("Surface Inventory")
    groups = sorted({g for w in result.windows for g in w.surface_by_group})
    rows = [["Cumulative mass on surface (t)"] + [w.window_id for w in result.windows]]
    for g in groups:
        rows.append([g] + [w.surface_by_group.get(g, 0.0) for w in result.windows])
    rows.append(["TOTAL hardware"] + [w.surface_hardware_t for w in result.windows])
    rows.append(["Avg surface load (kW)"] + [w.surface_avg_load_kw for w in result.windows])
    rows.append(["Installed generation (kWe)"] + [w.installed_generation_kwe for w in result.windows])
    rows.append(["Propellant banked (t)"] + [w.propellant_cumulative_t for w in result.windows])
    rows.append([""])
    rows.append(["Final inventory", "Qty", "Mass (t)"])
    for cid, qty, mass in result.windows[-1].surface_inventory:
        rows.append([cid, qty, mass])
    _sheet_rows(ws, rows)

    if result.violations:
        ws = wb.create_sheet("Violations")
        _sheet_rows(ws, [["Violation"]] + [[v] for v in result.violations])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
