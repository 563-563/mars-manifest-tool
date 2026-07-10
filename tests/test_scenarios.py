"""Scenario resolution + diffing (spec origin: HANDOFF.md §5.4), and report
smoke tests."""
import pytest

from mars_manifest import report as rpt
from mars_manifest.budgets import BudgetEngine
from mars_manifest.campaign import CampaignPlanner
from mars_manifest.packing import PackingEngine
from mars_manifest.scenarios import ScenarioError, compare


def test_scenario_presets_exist(manager):
    names = manager.names()
    for expected in ("baseline", "optimistic", "conservative", "solar_only"):
        assert expected in names


def test_override_resolution(manager):
    opt = manager.resolve("optimistic")
    assert opt.get("fleet.tankers_per_ship") == 10
    assert opt.get("cost.active_launch_cost") == "aspirational"
    assert opt.get("overheads.spares_fraction_of_dry") == 0.25
    # untouched values inherit from baseline
    assert opt.get("power.fission_unit_kwe") == 40


def test_unknown_scenario_raises(manager):
    with pytest.raises(ScenarioError):
        manager.resolve("does_not_exist")


def test_compare_mission(manager, catalog, precursor):
    comp = compare(manager, catalog, "optimistic", "conservative", mission=precursor)
    rows = {r.metric: r for r in comp.rows}
    assert rows["total launches"].a == 55   # 5 x (1+10)
    assert rows["total launches"].b == 85   # 5 x (1+16)
    assert rows["tankers per ship"].changed
    # conservative carries more spares -> heavier grand total
    assert rows["grand total mass (t)"].b > rows["grand total mass (t)"].a


def test_compare_campaign(manager, catalog, campaign_4w):
    comp = compare(manager, catalog, "baseline", "optimistic", campaign=campaign_4w)
    rows = {r.metric: r for r in comp.rows}
    assert rows["first crew window"].a == "2033-03"
    assert rows["total launches"].a > rows["total launches"].b


def test_reports_render(tmp_path, manager, catalog, precursor, campaign_4w):
    a = manager.resolve("baseline")
    budget = BudgetEngine(catalog, a).compute(precursor)
    packing = PackingEngine(catalog, a).pack(precursor, budget)

    md = rpt.budget_markdown(budget, precursor, packing)
    assert "GRAND TOTAL" in md and "Launch math" in md

    xlsx = rpt.budget_xlsx(budget, precursor, catalog, a, tmp_path / "budget.xlsx", packing)
    assert xlsx.exists()

    planner = CampaignPlanner(catalog, a, manager.capability_unlocks(), manager.crewed_requires())
    result = planner.run(campaign_4w)
    md = rpt.campaign_markdown(result)
    assert "First crew window" in md
    xlsx = rpt.campaign_xlsx(result, tmp_path / "campaign.xlsx")
    assert xlsx.exists()

    # rendered workbook numbers come from the engine, not recomputed
    from openpyxl import load_workbook
    wb = load_workbook(tmp_path / "budget.xlsx")
    summary = {r[0]: r[1] for r in wb["Summary"].iter_rows(values_only=True) if r[0]}
    assert summary["GRAND TOTAL MASS"] == pytest.approx(budget.mass.grand_total_t)


def test_spacex_internal_splits_cargo_and_tanker_rates(manager, catalog, precursor):
    from mars_manifest.packing import PackingEngine
    a = manager.resolve("spacex_internal")
    packed = PackingEngine(catalog, a).pack(precursor)
    lm = packed.launch
    assert lm.split_rates
    assert lm.cargo_ship_rate_musd == 30
    assert lm.tanker_rate_musd == 12
    # 5 expended cargo ships x $30M + 80 reused tanker flights x $12M
    assert lm.launch_cost_musd == pytest.approx(5 * 30 + 80 * 12)


def test_fsp_2025_directive_scenario(manager, catalog, precursor):
    from mars_manifest.budgets import BudgetEngine
    a = manager.resolve("fsp_2025")
    b = BudgetEngine(catalog, a).compute(precursor)
    # 100 kWe units: ceil(354.15/100) = 4 units at the 15 t lander ceiling
    assert b.fission.units == 4
    assert b.fission.mass_t == pytest.approx(60.0)
    # heavier gen than the 40 kWe path (54 t) -> grand total rises
    assert b.mass.grand_total_t == pytest.approx(245.1, abs=0.2)


def test_baseline_launch_math_unchanged_by_split_feature(catalog, baseline, precursor):
    from mars_manifest.packing import PackingEngine
    lm = PackingEngine(catalog, baseline).pack(precursor).launch
    assert not lm.split_rates
    assert lm.launch_cost_musd == pytest.approx(85 * 90.0)


def test_informed_spares_per_group(manager, catalog, precursor):
    from mars_manifest.budgets import BudgetEngine
    a = manager.resolve("informed_spares")
    b = BudgetEngine(catalog, a).compute(precursor)
    # hand-computed: group masses x literature fractions + 0.15 on gen/storage
    assert b.mass.spares_t == pytest.approx(37.5, abs=0.05)
    assert b.mass.grand_total_t == pytest.approx(220.8, abs=0.1)
    # baseline flat sparing must be untouched (regression contract)
    base = BudgetEngine(catalog, manager.resolve("baseline")).compute(precursor)
    assert base.mass.spares_t == pytest.approx(53.08, abs=0.05)


def test_ship_manifest_detail_sums_match(catalog, baseline, precursor):
    from mars_manifest.budgets import BudgetEngine
    from mars_manifest.packing import PackingEngine
    budget = BudgetEngine(catalog, baseline).compute(precursor)
    packed = PackingEngine(catalog, baseline).pack(precursor, budget)
    for s in packed.ships:
        assert sum(d[2] for d in s.manifest_detail) == pytest.approx(s.mass_t)
        assert sum(d[3] for d in s.manifest_detail) == pytest.approx(s.volume_m3)


def test_window_enablements_read_like_english(manager, catalog, campaign_4w):
    from mars_manifest.campaign import CampaignPlanner
    from mars_manifest.report import window_enablements
    a = manager.resolve("baseline")
    planner = CampaignPlanner(catalog, a, manager.capability_unlocks(), manager.crewed_requires())
    result = planner.run(campaign_4w)
    lines = window_enablements(result.windows[0], catalog, a)
    text = " ".join(lines)
    assert "kWe" in text and "methalox" in text and "ISS" in text
    crew_lines = window_enablements(result.windows[-1], catalog, a)
    assert any("First crew" in ln for ln in crew_lines)


def test_aiaa_contingency_per_group(manager, catalog, precursor):
    from mars_manifest.budgets import BudgetEngine
    base = BudgetEngine(catalog, manager.resolve("baseline")).compute(precursor)
    aiaa = BudgetEngine(catalog, manager.resolve("aiaa_contingency")).compute(precursor)
    # AIAA MGA (mostly 20-30%) is heavier than the flat 10% baseline
    assert aiaa.mass.contingency_t > base.mass.contingency_t
    assert aiaa.mass.grand_total_t > base.mass.grand_total_t
    # baseline flat contingency must be exactly 10% of hardware (regression)
    hw = base.mass.fixed_hardware_t + base.mass.generation_t + base.mass.storage_t
    assert base.mass.contingency_t == pytest.approx(0.10 * hw, abs=0.01)
